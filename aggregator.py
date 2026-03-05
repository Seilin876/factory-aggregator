import os
import sys
import pandas as pd
import configparser
import logging
import openpyxl
from datetime import datetime
import glob
import re
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# 初始化日誌記錄
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def create_summary_dashboard(writer, df, date_str):
    """
    建立 'Summary_Dashboard' 分頁。
    上半部：異常模式分析總表。
    下半部：各線別詳細報表。
    """
    workbook = writer.book
    sheet_name = 'Summary_Dashboard'
    
    # --- 1. 資料預處理 ---
    # A. 將關鍵欄位轉換為數值型態
    # 以 'dB(A)' 判斷部分異常狀況
    numeric_cols = [
        'index1', 'Index1_Limit', 
        'index2', 'Index2_Limit', 
        'index3', 'Index3_Limit', 
        'RPM', 'RPM_Low', 'RPM_Up', 'dB(A)'
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        else:
            df[col] = 0

    # B. 定義基本判斷條件
    cond_rotating = (df['RPM'] != 0)
    cond_idx1_high = (df['index1'] > df['Index1_Limit'])
    cond_idx2_high = (df['index2'] > df['Index2_Limit'])
    cond_idx3_high = (df['index3'] > df['Index3_Limit'])
    cond_out_of_control = (df['RPM'] > df['RPM_Low']) | (df['RPM'] > 10000)
    cond_no_rotate = (df['RPM'] == 0)

    # --- 2. 詳細項目計數邏輯 ---

    df['is_fail'] = df['Total_Result'].apply(lambda x: 0 if str(x).strip().upper() == 'OK' else 1)

    # 綜合 Noise 判定 (任一Index超標且風扇有運轉，或是頻譜超規格)
    df['calc_noise'] = (
        ((cond_idx1_high | cond_idx2_high | cond_idx3_high) & cond_rotating).astype(int) +
        df.apply(lambda row: 1 if str(row.get('Intelligent_Control','')).strip().upper() == 'OK' 
                 and str(row.get('Total_Result','')).strip().upper() != 'OK' else 0, axis=1)
    )

    # 各項 Index 獨立 Fail 判定
    df['calc_only_idx1'] = (cond_idx1_high & ~cond_idx2_high & ~cond_idx3_high & cond_rotating).astype(int)
    df['calc_only_idx2'] = (~cond_idx1_high & cond_idx2_high & ~cond_idx3_high & cond_rotating).astype(int)
    df['calc_only_idx3'] = (~cond_idx1_high & ~cond_idx2_high & cond_idx3_high & cond_rotating).astype(int)
    
    # 兩項(含)以上 Index 異常判定
    df['calc_multi_idx'] = (((cond_idx1_high.astype(int) + cond_idx2_high.astype(int) + cond_idx3_high.astype(int)) >= 2) & cond_rotating).astype(int)
    
    df['calc_spec_fail'] = df.apply(lambda row: 1 if str(row.get('Intelligent_Control','')).strip().upper() == 'OK' 
                                    and str(row.get('Total_Result','')).strip().upper() != 'OK' else 0, axis=1)
    
    df['calc_out_control'] = cond_out_of_control.astype(int)
    df['calc_no_rotate'] = cond_no_rotate.astype(int)
    df['calc_rpm_ng'] = ((cond_out_of_control) | (cond_no_rotate)).astype(int)

    df['calc_pause'] = df['Model_Name'].apply(lambda x: 1 if 'pauseorfreerun' in str(x).lower().replace(" ", "") else 0)
    df['calc_no_barcode'] = df['Barcode'].apply(lambda x: 1 if pd.isna(x) or str(x).strip() == '' else 0)
    df['calc_others'] = ((df['calc_pause'] == 1) | (df['calc_no_barcode'] == 1)).astype(int)


    # --- 3. 邏輯分析階段 (預先計算總表所需數據) ---
    
    lines = sorted(df['Line_Name'].unique())
    
    # 失效模式類別
    detected_failures = {
        'Carrier Abnormal': [],
        'Test Pin Abnormal': [],
        'Mic Position Variant': [],
        'Mic Cable Abnormal': [],
        'Isolation Box Abnormal': [],
        'Need to Check Audio File': []
    }

    # 輔助函式：將 Line_13 轉換為 C13，Station_1 轉換為 No1，方便TE閱讀Summary
    def format_location(line_str, station_str=None):
        # 1. 處理線別名稱
        l_nums = re.findall(r'\d+', str(line_str))
        l_code = f"C{int(l_nums[0]):02d}" if l_nums else str(line_str)
        
        # 2. 處理站點名稱
        if station_str:
            s_nums = re.findall(r'\d+', str(station_str))
            s_code = f"No{int(s_nums[0])}" if s_nums else str(station_str)
            return f"{l_code}-{s_code}"
        
        return l_code

    for line in lines:
        line_df = df[df['Line_Name'] == line]
        stations = sorted(line_df['Device_ID'].unique())
        
        # 輔助計算：統計該線別各站點的數據
        st_stats = {}
        for st in stations:
            st_data = line_df[line_df['Device_ID'] == st]
            total = len(st_data)
            
            rpm_fail_count = st_data['calc_rpm_ng'].sum()
            other_fail_count = st_data['calc_others'].sum()
            
            rpm_rate = (rpm_fail_count / total) if total > 0 else 0
            other_rate = (other_fail_count / total) if total > 0 else 0
            
            idx1_mean = st_data['index1'].mean() if total > 0 else 0
            idx2_mean = st_data['index2'].mean() if total > 0 else 0
            idx3_mean = st_data['index3'].mean() if total > 0 else 0
            dba_mean = st_data['dB(A)'].mean() if 'dB(A)' in st_data.columns and total > 0 else 0
            
            cable_fail_count = ((st_data['index1'] > 300) | (st_data['index2'] > 300) | (st_data['index3'] > 300)).sum()
            
            st_stats[st] = {
                'rpm_rate': rpm_rate,
                'other_rate': other_rate,
                'idx1_mean': idx1_mean,
                'idx2_mean': idx2_mean,
                'idx3_mean': idx3_mean,
                'dba_mean': dba_mean,
                'cable_fail_count': cable_fail_count
            }

        # 異常模式 1：載具異常
        # 條件：各站點因「轉速異常/其他異常拋料率」> 1% 且數值接近 (最大差距 <= 3%)
        if stations:
            rpm_rates = [st_stats[s]['rpm_rate'] for s in stations]
            other_rates = [st_stats[s]['other_rate'] for s in stations]

            all_rpm_high = all(r > 0.01 for r in rpm_rates)
            all_other_high = all(r > 0.01 for r in other_rates)

            rpm_consistent = (max(rpm_rates) - min(rpm_rates) <= 0.03) if rpm_rates else False
            other_consistent = (max(other_rates) - min(other_rates) <= 0.03) if other_rates else False
            
            if (all_rpm_high and rpm_consistent) or (all_other_high and other_consistent):
                short_code = format_location(line)
                detected_failures['Carrier Abnormal'].append(f"{short_code}-All")

        # 異常模式 2：探針異常
        # 條件：各站點因「轉速異常/其他異常拋料率」> 2%
        for s in stations:
            if (st_stats[s]['rpm_rate'] > 0.02) or (st_stats[s]['other_rate'] > 0.02):
                detected_failures['Test Pin Abnormal'].append(format_location(line, s))

        # 異常模式 3, 5, 6,多個站點比較
        if len(stations) > 1:
            for s in stations:
                other_stations = [os for os in stations if os != s]
                
                # 計算其他站點平均值
                others_idx1_mean = sum(st_stats[os]['idx1_mean'] for os in other_stations) / len(other_stations)
                others_idx2_mean = sum(st_stats[os]['idx2_mean'] for os in other_stations) / len(other_stations)
                others_idx3_mean = sum(st_stats[os]['idx3_mean'] for os in other_stations) / len(other_stations)
                
                # 邏輯 3：麥克風位置變異
                # 條件：單台機之index值<同條線其他設備0.8倍
                if (st_stats[s]['idx1_mean'] < others_idx1_mean * 0.8) or \
                   (st_stats[s]['idx2_mean'] < others_idx2_mean * 0.8) or \
                   (st_stats[s]['idx3_mean'] < others_idx3_mean * 0.8):
                    detected_failures['Mic Position Variant'].append(format_location(line, s))

                # 邏輯 5：隔音箱異常
                # 條件：單台機之dB值>同條線其他設備1.1倍
                others_dba_mean = sum(st_stats[os]['dba_mean'] for os in other_stations) / len(other_stations)
                if others_dba_mean > 0:
                    if st_stats[s]['dba_mean'] > others_dba_mean * 1.1:
                         detected_failures['Isolation Box Abnormal'].append(format_location(line, s))

                # 邏輯 6：需確認音頻
                # 條件：單台機之index值>同條線其他設備1.3倍
                if (st_stats[s]['idx1_mean'] > others_idx1_mean * 1.3) or \
                   (st_stats[s]['idx2_mean'] > others_idx2_mean * 1.3) or \
                   (st_stats[s]['idx3_mean'] > others_idx3_mean * 1.3):
                    detected_failures['Need to Check Audio File'].append(format_location(line, s))

        # 異常模式 4：線材異常
        # 條件：單台機之index值>300的次數超過10次
        for s in stations:
            if st_stats[s]['cable_fail_count'] > 10:
                detected_failures['Mic Cable Abnormal'].append(format_location(line, s))


    # --- 4. 生成報表 ---
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
    ws = workbook[sheet_name]

    current_row = 1
    
    # 設定儲存格樣式
    thin_border = Border(left=Side(style='thin', color='FFFFFF'), right=Side(style='thin', color='FFFFFF'), top=Side(style='thin', color='FFFFFF'), bottom=Side(style='thin', color='FFFFFF'))
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Summary樣式
    fail_header_fill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid') 
    fail_header_font = Font(bold=True, color='000000', size=14)
    fail_row_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid') 
    fail_row_font = Font(bold=False, color='000000')

    # 各線別詳細報表樣式
    fill_g1_main = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid') 
    font_g1_main = Font(bold=True, color='000000')
    fill_g1_sub = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') 
    font_sub_default = Font(bold=False, color='000000')

    fill_g2_main = PatternFill(start_color='BDE7FF', end_color='BDE7FF', fill_type='solid')
    font_g2_main = Font(bold=True, color='000000')
    fill_g2_sub = PatternFill(start_color='D1EFFF', end_color='D1EFFF', fill_type='solid')

    fill_g3_main = PatternFill(start_color='A0E8E6', end_color='A0E8E6', fill_type='solid')
    font_g3_main = Font(bold=True, color='000000')
    fill_g3_sub = PatternFill(start_color='C6F1F0', end_color='C6F1F0', fill_type='solid')

    fill_g4_main = PatternFill(start_color='C9F084', end_color='C9F084', fill_type='solid')
    font_g4_main = Font(bold=True, color='000000')
    fill_g4_sub = PatternFill(start_color='E0F6B8', end_color='E0F6B8', fill_type='solid')

    # 排行榜顏色
    rank_1_fill = PatternFill(start_color='FF5050', end_color='FF5050', fill_type='solid')
    rank_2_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
    rank_3_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    rank_4_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')


    # ==========================================
    # 第一部分：寫入異常模式總表
    # ==========================================
    
    # 1. 寫入標題
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    title_cell = ws.cell(row=current_row, column=1, value=f"Daily Failure Mode Analysis ({date_str})")
    title_cell.font = fail_header_font
    title_cell.fill = fail_header_fill
    title_cell.alignment = center_align
    current_row += 1

    # 2. 寫入欄位標題
    ws.cell(row=current_row, column=1, value="No.").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value="Failure Mode").font = Font(bold=True)
    ws.cell(row=current_row, column=3, value="Detected Locations (Line - Station)").font = Font(bold=True)
    
    for col in range(1, 4):
        c = ws.cell(row=current_row, column=col)
        c.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid') 
        c.border = thin_border
        c.alignment = center_align
    current_row += 1

    # 3. 依序列出 6 種失效模式
    modes_order = [
        'Carrier Abnormal', 
        'Test Pin Abnormal', 
        'Mic Position Variant', 
        'Mic Cable Abnormal', 
        'Isolation Box Abnormal', 
        'Need to Check Audio File'
    ]

    for idx, mode in enumerate(modes_order, 1):
        # 序號
        c1 = ws.cell(row=current_row, column=1, value=idx)
        c1.fill = fail_row_fill
        c1.border = thin_border
        c1.alignment = center_align

        # 失效模式名稱
        c2 = ws.cell(row=current_row, column=2, value=mode)
        c2.fill = fail_row_fill
        c2.border = thin_border
        c2.alignment = Alignment(horizontal='left', vertical='center')

        # 偵測結果
        locations = detected_failures.get(mode, [])
        if locations:
            res_str = ", ".join(locations)
            # 若有發現異常，列出發生位置，將字體顯示紅色
            font_res = Font(color='FF0000', bold=True)
        else:
            res_str = "OK"
            # 正常則顯示OK(綠色)
            font_res = Font(color='008000', bold=True)

        c3 = ws.cell(row=current_row, column=3, value=res_str)
        c3.fill = fail_row_fill
        c3.border = thin_border
        c3.font = font_res
        c3.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        current_row += 1

    # 總表與詳細報表之間加入空白行
    current_row += 2 


    # ==========================================
    # 第二部分：寫入各線別詳細報表
    # ==========================================
    
    for line in lines:
        line_df = df[df['Line_Name'] == line]
        stations = sorted(line_df['Device_ID'].unique())
        
        # 標題：線別名稱
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(stations)+2)
        cell = ws.cell(row=current_row, column=1, value=line)
        cell.font = Font(bold=True, size=14, color='000000')
        cell.fill = fill_g1_main
        cell.alignment = center_align
        current_row += 1

        # 欄位標題
        headers = ['Metric', 'Total'] + stations
        for i, h in enumerate(headers):
            c = ws.cell(row=current_row, column=i+1, value=h)
            c.font = Font(bold=True)
            c.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
            c.border = thin_border
            c.alignment = center_align
        current_row += 1

        # 指標配置
        metrics_config = [
            # Group 1
            ('Total Count',       None,               False, False, 1, False),
            ('Fail Count',        'is_fail',          False, False, 1, False),
            ('Fail Rate',         'is_fail',          True,  False, 1, False),
            ('Noise Rate',        'calc_noise',       True,  False, 1, False),
            ('RPM Fail Rate',     'calc_rpm_ng',      True,  False, 1, False),
            ('Other Fail Rate',   'calc_others',      True,  False, 1, False),
            
            # Group 2
            ('Noise',             'calc_noise',       False, False, 2, True),
            ('Only Index1 Fail',  'calc_only_idx1',   False, False, 2, False),
            ('Only Index2 Fail',  'calc_only_idx2',   False, False, 2, False),
            ('Only Index3 Fail',  'calc_only_idx3',   False, False, 2, False),
            ('Multiple Index Fail','calc_multi_idx',  False, False, 2, False),
            ('Spec Fail',         'calc_spec_fail',   False, False, 2, False),

            # Group 3
            ('RPM NG',            'calc_rpm_ng',      False, False, 3, True),
            ('Out of control',    'calc_out_control', False, False, 3, False),
            ('No rotate',         'calc_no_rotate',   False, False, 3, False),

            # Group 4
            ('Others',            'calc_others',      False, False, 4, True),
            ('PauseOrFreeRun',    'calc_pause',       False, False, 4, False),
            ('No Barcode',        'calc_no_barcode',  False, False, 4, False),
            ('Model Name',        'Model_Name',       False, True,  4, False),
        ]

        ranking_targets = ['Fail Count', 'Fail Rate', 'Noise Rate', 'RPM Fail Rate', 'Other Fail Rate']

        for label, col, is_rate, is_string, group_id, is_main in metrics_config:
            
            # 判斷所屬 Group 並套用對應底色
            if group_id == 1:
                base_fill = fill_g1_main if is_main else fill_g1_sub
                base_font = font_g1_main if is_main else font_sub_default
            elif group_id == 2:
                base_fill = fill_g2_main if is_main else fill_g2_sub
                base_font = font_g2_main if is_main else font_sub_default
            elif group_id == 3:
                base_fill = fill_g3_main if is_main else fill_g3_sub
                base_font = font_g3_main if is_main else font_sub_default
            else: # Group 4
                base_fill = fill_g4_main if is_main else fill_g4_sub
                base_font = font_g4_main if is_main else font_sub_default

            # 寫入指標名稱
            c_label = ws.cell(row=current_row, column=1, value=label)
            c_label.fill = base_fill
            c_label.font = base_font
            c_label.border = thin_border

            # "Model Name" Row合併儲存格，並隱藏"Model Name"有"pauseorfreerun"的情形:
            if label == 'Model Name':
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=2+len(stations))
                
                unique_names = line_df[col].dropna().unique()
                valid_names = [str(n) for n in unique_names if 'pauseorfreerun' not in str(n).lower().replace(" ", "")]
                val = ",".join(valid_names)
                
                c_total = ws.cell(row=current_row, column=2, value=val)
                c_total.fill = base_fill
                c_total.font = base_font
                c_total.border = thin_border
                c_total.alignment = center_align
                
                current_row += 1
                continue
            
            # 計算數據
            station_values = []
            
            # 單線整體(Total)欄位計算
            if is_string:
                unique_names = line_df[col].dropna().unique()
                valid_names = [str(n) for n in unique_names if 'pauseorfreerun' not in str(n).lower().replace(" ", "")]
                val = ",".join(valid_names)
            elif label == 'Total Count':
                val = len(line_df)
            elif is_rate:
                total = len(line_df)
                fails = line_df[col].sum()
                val = f"{(fails/total)*100:.2f}%" if total > 0 else "0.00%"
            else:
                val = line_df[col].sum()
            
            c_total = ws.cell(row=current_row, column=2, value=val)
            c_total.fill = base_fill
            c_total.font = base_font
            c_total.border = thin_border
            c_total.alignment = center_align

            # 各台機欄位計算
            for i, station in enumerate(stations):
                st_df = line_df[line_df['Device_ID'] == station]
                
                raw_num_val = 0
                if is_string:
                    unique_names = st_df[col].dropna().unique()
                    valid_names = [str(n) for n in unique_names if 'pauseorfreerun' not in str(n).lower().replace(" ", "")]
                    st_val = valid_names[0] if len(valid_names) > 0 else ""
                elif label == 'Total Count':
                    raw_num_val = len(st_df)
                    st_val = raw_num_val
                elif is_rate:
                    total = len(st_df)
                    fails = st_df[col].sum()
                    raw_num_val = (fails/total) if total > 0 else 0
                    st_val = f"{raw_num_val*100:.2f}%"
                else:
                    raw_num_val = st_df[col].sum()
                    st_val = raw_num_val
                
                c_st = ws.cell(row=current_row, column=i+3, value=st_val)
                c_st.fill = base_fill 
                c_st.font = base_font
                c_st.border = thin_border
                c_st.alignment = center_align
                
                station_values.append({'col_idx': i+3, 'val': raw_num_val, 'cell': c_st})

            # 排名上色邏輯
            if label in ranking_targets:
                unique_vals = sorted(list(set([x['val'] for x in station_values])), reverse=True)
                
                val_1st = unique_vals[0] if len(unique_vals) > 0 else None
                val_2nd = unique_vals[1] if len(unique_vals) > 1 else None
                val_3rd = unique_vals[2] if len(unique_vals) > 2 else None
                val_4rd = unique_vals[3] if len(unique_vals) > 3 else None

                for item in station_values:
                    if item['val'] == 0: continue
                        
                    if item['val'] == val_1st:
                        item['cell'].fill = rank_1_fill 
                    elif item['val'] == val_2nd:
                        item['cell'].fill = rank_2_fill 
                    elif item['val'] == val_3rd:
                        item['cell'].fill = rank_3_fill 
                    elif item['val'] == val_4rd:
                        item['cell'].fill = rank_4_fill

            current_row += 1
        
        current_row += 1 # 各線別之間的空白行

def run_aggregation(target_date=None):
    """
    整合各機台產出的測試日誌檔並生成報表。
    """
    if getattr(sys, 'frozen', False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    
    config_path = os.path.join(base_dir, 'config.ini')

    if not os.path.exists(config_path):
        logging.error(f"找不到 Config 檔案路徑: {config_path}")
        return

    config = configparser.ConfigParser()
    config.read(config_path, encoding='utf-8')

    try:
        source_dir = config['Path']['Source_Folder']
        output_base = config['Path']['Output_Folder']
        if output_base.startswith('.\\') or output_base.startswith('./'):
            output_dir = os.path.join(base_dir, output_base.replace('.\\', '').replace('./', ''))
        else:
            output_dir = output_base
    except KeyError as e:
        logging.error(f"Config 檔案缺少鍵值: {e}")
        return
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    if target_date is None:
        from datetime import timedelta
        target_date = (datetime.now() - timedelta(days=1)).strftime('%Y%m%d')
    
    logging.info(f"開始整合日期: {target_date} 之資料")

    device_map = {}
    if 'Device_Mapping' in config:
        for ip, mapping in config['Device_Mapping'].items():
            try:
                line, station = mapping.split(',')                
                formatted_station = station.strip().replace('_', ' ')
                device_map[ip.replace('.', '_')] = {'Line': line.strip(), 'Station': formatted_station}
            except ValueError:
                logging.error(f"解析 IP {ip} 對應之格式錯誤: {mapping}")

    search_pattern = os.path.join(source_dir, f"{target_date}_*.txt")
    files = glob.glob(search_pattern)
    
    if not files:
        logging.warning(f"在 {source_dir} 找不到日期 {target_date} 的日誌檔")
        return

    all_data = []

    for file_path in files:
        filename = os.path.basename(file_path)
        match = re.match(rf"{target_date}_(.+)\.txt", filename)
        if not match:
            continue
            
        ip_key = match.group(1)
        meta = device_map.get(ip_key, {'Line': 'Unknown_Line', 'Station': f'Unknown {ip_key}'})

        try:
            df = pd.read_csv(file_path, sep='\t', encoding='cp950', on_bad_lines='skip', index_col=False)
            df.columns = df.columns.str.strip()
            
            # 修正舊版軟體的欄位拼寫錯誤
            if 'Toral_Result' in df.columns:
                df.rename(columns={'Toral_Result':'Total_Result'}, inplace=True)
                logging.info(f"已修正檔案 {filename} 中的錯字 'Toral_Result'")

            if 'Total_Result' not in df.columns:
                logging.warning(f"跳過檔案 {filename}: 缺少 'Total_Result' 欄位")
                continue

            df['Line_Name'] = meta['Line']
            df['Device_ID'] = meta['Station']
            df['Source_IP'] = ip_key.replace('_', '.')
            df['Log_Date'] = target_date
            df['Total_Result'] = df['Total_Result'].astype(str)
            
            all_data.append(df)
            logging.info(f"已處理檔案: {filename} (共 {len(df)} 筆資料)")

        except UnicodeDecodeError:
            try:
                df = pd.read_csv(file_path, sep='\t', encoding='utf-8', on_bad_lines='skip', index_col=False)
                df.columns = df.columns.str.strip()
                df['Line_Name'] = meta['Line']
                df['Device_ID'] = meta['Station']
                df['Source_IP'] = ip_key.replace('_', '.')
                df['Log_Date'] = target_date
                df['Total_Result'] = df['Total_Result'].astype(str)
                all_data.append(df)
                logging.info(f"使用 UTF-8 編碼成功處理檔案: {filename}")
            except Exception as e2:
                logging.error(f"嘗試 UTF-8 編碼讀取檔案 {filename} 失敗: {e2}")

        except Exception as e:
            logging.error(f"讀取檔案 {filename} 發生未知錯誤: {e}")

    if all_data:
        master_df = pd.concat(all_data, ignore_index=True)
        output_file = os.path.join(output_dir, f"Daily_Summary_{target_date}.xlsx")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            master_df.to_excel(writer, sheet_name=target_date, index=False)
            create_summary_dashboard(writer, master_df, target_date)
            
        logging.info(f"資料整合完畢。輸出檔案為: {output_file}")
    else:
        logging.warning("未找到有效資料，無法產出報表。")

if __name__ == "__main__":
    while True:
        # 詢問 user 日期
        user_date= input("請輸入要執行的日期(格式為YYYYMMDD,例如 20260101):").strip()
        # 簡單驗證輸入格式是否為 8 個數字
        if re.match(r"^\d{8}$", user_date):
            print(f"準備執行日期 {user_date} 的測試紀錄...")
            run_aggregation(user_date)
            break # 執行完畢後跳出迴圈
        else:
            print("輸入格式錯誤!請重新輸入 8 位數字的日期格式(例如:20260101)。\n")
